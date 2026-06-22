# charts.py — Spreadsheet chart generation via openpyxl.
# SPDX-License-Identifier: GPL-3.0-or-later

"""Create charts from spreadsheet data and embed them in XLSX files.

Uses openpyxl's chart API (bar, line, pie, scatter).  Charts are
added at save time so they appear when the file is opened in
LibreOffice or Excel.
"""

from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter


CHART_TYPES = {
    'bar': BarChart,
    'line': LineChart,
    'pie': PieChart,
}


def add_chart_to_workbook(wb, sheet_name, data_range, chart_type='bar',
                          title=None, position='E2'):
    """Add a chart to an openpyxl Workbook.

    Args:
        wb: openpyxl Workbook (with data already written).
        sheet_name: Name of the worksheet containing the data.
        data_range: Tuple of (start_row, start_col, end_row, end_col)
                    as 1-based indices (e.g. (1, 1, 5, 2) for A1:B5).
        chart_type: 'bar', 'line', or 'pie'.
        title: Optional chart title string.
        position: Cell reference for top-left corner of the chart.

    Returns:
        The openpyxl Chart object (appended to the sheet).
    """
    chart_class = CHART_TYPES.get(chart_type, BarChart)
    chart = chart_class()
    chart.style = 10

    if title:
        chart.title = title

    ws = wb[sheet_name]
    r1, c1, r2, c2 = data_range

    data = Reference(ws, min_col=c1, min_row=r1, max_col=c2, max_row=r2)
    cats = Reference(ws, min_col=c1, min_row=r1 + 1, max_row=r2)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # Set axis labels from first row (bar/line only; pie has no axes)
    if chart_type != 'pie' and c2 > c1:
        chart.y_axis.title = 'Values'
        chart.x_axis.title = ws.cell(row=r1, column=c1).value or ''

    ws.add_chart(chart, position)
    return chart


def add_chart_to_file(xlsx_path, sheet_name, data_range, chart_type='bar',
                      title=None, position='E2'):
    """Open an existing XLSX, add a chart, and save.

    Convenience wrapper for post-processing saved files.
    """
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    add_chart_to_workbook(wb, sheet_name, data_range, chart_type, title, position)
    wb.save(xlsx_path)
